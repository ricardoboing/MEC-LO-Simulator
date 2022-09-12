from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

def _get_cell_name(rowIndex, columnIndex):
	rowIndex = str(rowIndex)
	columnLetter = get_column_letter(columnIndex)

	return columnLetter + rowIndex

class WorkbookTableFile:
	def __init__(self, fileName, currentSheetName=""):
		self.file = load_workbook(fileName)
		self.set_current_sheet(currentSheetName)

	def set_current_sheet(self, currentSheetName):
		if currentSheetName == "":
			self.currentSheet = self.file.active
		else:
			self.currentSheet = self.file[currentSheetName]

	def get_column_name(self, columnIndex):
		return self.get_cell_value(0, columnIndex)

	def get_row_name(self, rowIndex):
		return self.get_cell_value(rowIndex, 0)

	def get_cell_value(self, rowIndex, columnIndex):
		columnIndex += 1
		rowIndex += 1

		cellName = _get_cell_name(rowIndex, columnIndex)
		value = self.currentSheet[cellName].value

		if value == "" or value == None:
			return None

		return value