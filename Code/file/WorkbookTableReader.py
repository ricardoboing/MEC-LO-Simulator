from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

def _get_cell_name(rowIndex, columnIndex):
	rowIndex = str(rowIndex)
	columnLetter = get_column_letter(columnIndex)

	return columnLetter + rowIndex

class WorkbookTableReader:
	def __init__(self, fileName, currentSheetName=""):
		self.file = load_workbook(fileName)
		self.set_current_sheet(currentSheetName)

	def set_current_sheet(self, currentSheetName):
		if currentSheetName == "":
			self.currentSheet = self.file.active
		else:
			self.currentSheet = self.file[currentSheetName]

	def get_cell_value(self, rowIndex, columnIndex):
		columnIndex += 1
		rowIndex += 1

		cellName = _get_cell_name(rowIndex, columnIndex)
		value = self.currentSheet[cellName].value

		if value == "" or value == None:
			return None

		return value

	def _get_column_name(self, columnIndex):
		return self.get_cell_value(0, columnIndex)

	def _get_row_name(self, rowIndex):
		return self.get_cell_value(rowIndex, 0)

	def _get_row_dict(self, rowIndex):
		rowDict = {}
		columnIndex = 1
		
		while True:
			columnName = self._get_column_name(columnIndex)

			if columnName == None:
				break

			rowDict[columnName] = self.get_cell_value(rowIndex, columnIndex)
			columnIndex += 1

		return rowDict

	def get_row_dict_list(self):
		rowDictList = {}
		rowIndex = 1

		while True:
			rowName = self._get_row_name(rowIndex)

			if rowName == None:
				break

			rowDictList[rowName] = self._get_row_dict(rowIndex)
			rowIndex += 1

		return rowDictList