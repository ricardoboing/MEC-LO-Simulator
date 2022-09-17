from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

def _get_cell_name(rowIndex, columnIndex):
	rowIndex = str(rowIndex)
	columnLetter = get_column_letter(columnIndex)

	return columnLetter + rowIndex

class WorkbookTableWriter:
	def __init__(self, fileName):
		self.fileName = fileName
		self.file = Workbook()
		
		for sheet in self.file:
			self.file.remove(sheet)

	def create_sheet(self, sheetName):
		self.file.create_sheet(title=sheetName)
		return self.file[sheetName]

	def save(self):
		self.file.save(self.fileName)

	@staticmethod
	def set_cell_value(sheet, rowIndex, columnIndex, value):
		rowIndex += 1
		columnIndex += 1

		cellName = _get_cell_name(rowIndex, columnIndex)
		sheet[cellName].value = value